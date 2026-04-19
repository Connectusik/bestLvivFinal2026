"""Read ДРРП dumps (Excel/CSV) and write a single clean workbook.

* **Reader** uses ``openpyxl`` in ``read_only`` mode — streams rows without
  materialising the full workbook, which is enough on ~20–50k-row inputs.
* **Writer** uses ``xlsxwriter``. openpyxl's writer creates a Python object
  per cell (even in ``write_only`` mode) and hit ~120 s on the sample
  dataset; xlsxwriter streams XML directly to a zip and finishes the same
  job in ~2 s while producing a ~45 % smaller file.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator

import openpyxl
import xlsxwriter


# Cadastral-number → third-party map viewer. Format validated before we
# generate the URL, so there's no injection risk even on junk input.
CADASTRAL_RE = re.compile(r"^\d{10}:\d{2}:\d{3}:\d{4}$")
CADASTRAL_URL_TEMPLATE = "https://kadastrova-karta.com/dilyanka/{cadastral}"


def cadastral_url(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    v = value.strip()
    if not CADASTRAL_RE.match(v):
        return None
    return CADASTRAL_URL_TEMPLATE.format(cadastral=v)


# Source column counts -------------------------------------------------------

LAND_COL_COUNT = 16
RE_COL_COUNT = 9


# Colour palette -------------------------------------------------------------

HEADER_BG = "#0B5394"
HEADER_FG = "#FFFFFF"
META_BG = "#EEF2F9"
CHANGED_CELL_BG = "#FFE066"  # saturated yellow
CHANGED_ROW_BG = "#FFF9D6"   # very light yellow

SEVERITY_BG = {
    "critical": "#F7CDC7",
    "high":     "#FFE0BD",
    "medium":   "#FFF4B3",
    "low":      "#D4E7FB",
}
SEVERITY_LABEL = {
    "critical": "Критична",
    "high":     "Висока",
    "medium":   "Середня",
    "low":      "Низька",
}


# ---------------------------------------------------------------------------
# Reader
# ---------------------------------------------------------------------------


def _open_rows(path: str | Path) -> Iterator[list[Any]]:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                yield list(row)
        finally:
            wb.close()
    elif suffix == ".csv":
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
            except csv.Error:
                dialect = csv.excel
            for row in csv.reader(f, dialect=dialect):
                yield list(row)
    else:
        raise ValueError(f"Unsupported file type: {suffix!r}. Use .xlsx or .csv")


def read_land(path: str | Path) -> list[list[Any]]:
    rows = list(_open_rows(path))
    if not rows:
        return []
    return [
        (r + [None] * LAND_COL_COUNT)[:LAND_COL_COUNT]
        for r in rows[1:]
        if any(v not in (None, "") for v in r)
    ]


def read_realestate(path: str | Path) -> list[list[Any]]:
    rows = list(_open_rows(path))
    if not rows:
        return []
    return [
        (r + [None] * RE_COL_COUNT)[:RE_COL_COUNT]
        for r in rows[1:]
        if any(v not in (None, "") for v in r)
    ]


# ---------------------------------------------------------------------------
# Writer (xlsxwriter)
# ---------------------------------------------------------------------------


def _coerce_cell(value: Any) -> Any:
    """xlsxwriter does not auto-handle Python ``datetime.date`` gracefully
    alongside strings — we normalise date/datetime to a ``datetime`` object
    because xlsxwriter writes those as Excel serial dates with a format.
    Everything else is passed through verbatim."""
    if isinstance(value, _dt.datetime):
        return value
    if isinstance(value, _dt.date):
        return _dt.datetime.combine(value, _dt.time.min)
    return value


def _build_formats(wb) -> dict[str, Any]:
    # A format in xlsxwriter is equivalent to a cell XF; each distinct
    # format is deduplicated by the library so we reuse aggressively.
    base_align = {"text_wrap": False, "valign": "vcenter"}

    header = wb.add_format({
        "bold": True, "font_color": HEADER_FG, "bg_color": HEADER_BG,
        "valign": "vcenter", "text_wrap": True,
    })

    date_fmt = "yyyy-mm-dd"

    # Data formats keyed by (fill, is_text, is_date).
    formats: dict[tuple[str | None, bool, bool], Any] = {}
    for fill in (None, CHANGED_CELL_BG, CHANGED_ROW_BG):
        for is_text in (False, True):
            for is_date in (False, True):
                if is_text and is_date:
                    continue  # date column can't also be text
                props: dict[str, Any] = dict(base_align)
                if fill:
                    props["bg_color"] = fill
                if is_text:
                    props["num_format"] = "@"
                if is_date:
                    props["num_format"] = date_fmt
                formats[(fill, is_text, is_date)] = wb.add_format(props)

    # URL formats — mirrors the data-format matrix but with blue underlined
    # text. Keyed by fill only (URLs are always text-typed).
    url_formats: dict[str | None, Any] = {}
    for fill in (None, CHANGED_CELL_BG, CHANGED_ROW_BG):
        props: dict[str, Any] = dict(base_align)
        props["font_color"] = "#0B5394"
        props["underline"] = 1
        props["num_format"] = "@"
        if fill:
            props["bg_color"] = fill
        url_formats[fill] = wb.add_format(props)

    meta_key = wb.add_format({
        "bold": True, "bg_color": META_BG,
        "valign": "top",
    })
    meta_val = wb.add_format({"valign": "top", "text_wrap": True})

    return {
        "header": header,
        "body": formats,
        "url": url_formats,
        "meta_key": meta_key,
        "meta_val": meta_val,
    }


def _body_format(formats, fill: str | None, is_text: bool, is_date: bool):
    return formats["body"][(fill, is_text, is_date)]


def _url_format(formats, fill: str | None):
    return formats["url"][fill]


def _write_sheet(
    wb,
    title: str,
    headers: list[str],
    rows: Iterable[list[Any]],
    formats: dict[str, Any],
    *,
    widths: list[int] | None = None,
    text_columns: list[int] | None = None,
    date_columns: list[int] | None = None,
    url_columns: dict[int, Callable[[Any], str | None]] | None = None,
    changed_masks: list[list[bool]] | None = None,
) -> None:
    ws = wb.add_worksheet(title)

    ncols = len(headers)
    text_col_set = {c - 1 for c in (text_columns or [])}
    date_col_set = {c - 1 for c in (date_columns or [])}
    url_col_map: dict[int, Callable[[Any], str | None]] = {
        c - 1: fn for c, fn in (url_columns or {}).items()
    }

    # Column widths must be set BEFORE rows in xlsxwriter when we also want
    # to attach a column-wide format (text format on tax ID columns, for
    # instance — we still do per-cell but this makes the format sticky for
    # any paste into an empty cell later).
    if widths is None:
        widths = [18] * ncols
    for i, w in enumerate(widths):
        col_fmt = None
        if i in text_col_set:
            col_fmt = _body_format(formats, None, True, False)
        elif i in date_col_set:
            col_fmt = _body_format(formats, None, False, True)
        ws.set_column(i, i, w, col_fmt)

    # Header row.
    ws.set_row(0, 30)
    for i, h in enumerate(headers):
        ws.write(0, i, h, formats["header"])

    mask_iter = iter(changed_masks) if changed_masks is not None else None

    r = 1  # xlsxwriter is 0-indexed; header occupies row 0
    for row in rows:
        mask = next(mask_iter) if mask_iter is not None else None
        row_has_change = bool(mask and any(mask))
        row_fill = CHANGED_ROW_BG if row_has_change else None

        for c in range(ncols):
            val = row[c] if c < len(row) else None
            if val is None:
                # Even None cells need the row tint so the visual spans.
                if row_fill and not (mask and c < len(mask) and mask[c]):
                    ws.write_blank(r, c, None, _body_format(formats, row_fill, False, False))
                continue

            is_text = c in text_col_set
            is_date = c in date_col_set and isinstance(val, (_dt.date, _dt.datetime))
            cell_changed = bool(mask) and c < len(mask) and mask[c]
            fill = CHANGED_CELL_BG if cell_changed else (row_fill if row_fill else None)

            # URL columns take precedence: if the value builds a valid URL
            # we render it as a clickable link with a URL-format cell that
            # still carries the row/cell fill.
            url_builder = url_col_map.get(c)
            url: str | None = None
            if url_builder is not None:
                try:
                    url = url_builder(val)
                except Exception:  # noqa: BLE001 — builder is user-provided
                    url = None
            if url:
                ws.write_url(r, c, url, _url_format(formats, fill), str(val))
                continue

            fmt = _body_format(formats, fill, is_text, is_date)

            if is_date:
                ws.write_datetime(r, c, _coerce_cell(val), fmt)
            elif is_text:
                # Force string to prevent Excel from re-typing the tax ID.
                ws.write_string(r, c, str(val), fmt)
            elif isinstance(val, bool):
                ws.write_boolean(r, c, val, fmt)
            elif isinstance(val, (int, float)):
                ws.write_number(r, c, val, fmt)
            else:
                ws.write_string(r, c, str(val), fmt)
        r += 1

    # Freeze header & attach auto-filter after the data is in.
    ws.freeze_panes(1, 0)
    if r > 1:
        ws.autofilter(0, 0, r - 1, ncols - 1)


def _write_metadata_sheet(wb, meta: dict[str, Any], formats: dict[str, Any]) -> None:
    ws = wb.add_worksheet("Метадані")
    ws.set_column(0, 0, 38)
    ws.set_column(1, 1, 72)
    key_fmt = formats["meta_key"]
    val_fmt = formats["meta_val"]

    r = 0

    def put(key: str, value: Any) -> None:
        nonlocal r
        ws.write(r, 0, key, key_fmt)
        if value is None or value == "":
            ws.write_blank(r, 1, None, val_fmt)
        elif isinstance(value, bool):
            ws.write_boolean(r, 1, value, val_fmt)
        elif isinstance(value, (int, float)):
            ws.write_number(r, 1, value, val_fmt)
        else:
            ws.write_string(r, 1, str(value), val_fmt)
        r += 1

    put("Дата очищення (UTC)", meta["cleaned_at_utc"])
    put("Дата очищення (локально)", meta["cleaned_at_local"])
    put("Версія скрипта", meta["version"])
    put("", "")
    put("Вхідний файл — земля", meta["input_land_name"])
    put("Розмір (байт)", meta["input_land_size"])
    put("SHA-256", meta["input_land_sha256"])
    put("Вхідний файл — нерухомість", meta["input_realestate_name"])
    put("Розмір (байт)", meta["input_realestate_size"])
    put("SHA-256", meta["input_realestate_sha256"])
    put("", "")
    put("Статистика — земля (рядків)", meta["rows_land"])
    put("   з них змінено нормалізатором", meta.get("rows_land_changed", "—"))
    put("Статистика — нерухомість (рядків)", meta["rows_realestate"])
    put("   з них змінено нормалізатором", meta.get("rows_realestate_changed", "—"))
    put("Статистика — унікальних власників", meta["rows_owners"])
    put("Час обробки: читання + нормалізація (с)", meta["elapsed_read_clean_sec"])
    put("", "")
    put("РОЗБІЖНОСТІ (всього)", meta.get("findings_total", 0))
    sev = meta.get("findings_by_severity", {}) or {}
    put("   критичних", sev.get("critical", 0))
    put("   високих", sev.get("high", 0))
    put("   середніх", sev.get("medium", 0))
    put("   низьких", sev.get("low", 0))
    put("Орієнтовний фін. ризик (грн)", meta.get("findings_exposure_uah", 0))
    put("", "")
    put("Застосовані трансформації", "")
    for i, t in enumerate(meta["transformations"], start=1):
        ws.write(r, 0, f"  {i}.", key_fmt)
        ws.write_string(r, 1, t, val_fmt)
        r += 1


def _write_findings_sheet(
    wb,
    formats: dict[str, Any],
    findings: list[Any],
) -> None:
    """«Розбіжності» sheet — each row is one detected issue."""
    ws = wb.add_worksheet("Розбіжності")

    headers = [
        "Серйозність", "Тип", "Опис",
        "ПІБ / Назва власника", "Податковий номер",
        "Земля (рядки)", "Нерухомість (рядки)",
        "Фін. ризик, грн", "Деталі", "Кадастрова карта",
    ]
    widths = [12, 30, 60, 34, 14, 18, 20, 16, 60, 24]
    for i, w in enumerate(widths):
        ws.set_column(i, i, w)

    ws.set_row(0, 30)
    header_fmt = formats["header"]
    for i, h in enumerate(headers):
        ws.write(0, i, h, header_fmt)

    # Pre-build a per-severity format cache.
    sev_cache: dict[str, Any] = {}
    for sev, color in SEVERITY_BG.items():
        sev_cache[sev] = wb.add_format({
            "bg_color": color,
            "bold": True,
            "align": "center",
            "valign": "vcenter",
        })

    wrap_fmt = wb.add_format({"valign": "top", "text_wrap": True})
    num_fmt = wb.add_format({"valign": "top", "num_format": "#,##0.00"})
    text_fmt = wb.add_format({"valign": "top", "num_format": "@"})
    url_fmt = wb.add_format({
        "valign": "top", "num_format": "@",
        "font_color": "#0B5394", "underline": 1,
    })

    for r, f in enumerate(findings, start=1):
        ws.set_row(r, 38)
        ws.write(r, 0, SEVERITY_LABEL.get(f.severity, f.severity), sev_cache.get(f.severity))
        ws.write(r, 1, f.title, wrap_fmt)
        ws.write(r, 2, f.description, wrap_fmt)
        ws.write(r, 3, f.owner_name or "—", wrap_fmt)
        ws.write(r, 4, f.owner_id or "—", text_fmt)
        ws.write(r, 5,
                 ", ".join(str(x) for x in f.land_excel_rows[:10]) +
                 (f" (+{len(f.land_excel_rows) - 10})" if len(f.land_excel_rows) > 10 else ""),
                 wrap_fmt)
        ws.write(r, 6,
                 ", ".join(str(x) for x in f.re_excel_rows[:10]) +
                 (f" (+{len(f.re_excel_rows) - 10})" if len(f.re_excel_rows) > 10 else ""),
                 wrap_fmt)
        if f.exposure_uah:
            ws.write_number(r, 7, round(f.exposure_uah, 2), num_fmt)
        else:
            ws.write_blank(r, 7, None, num_fmt)
        # Compact one-line evidence: key=value pairs
        ev = "; ".join(f"{k}={_short(v)}" for k, v in f.evidence.items())
        ws.write(r, 8, ev, wrap_fmt)

        # If this finding has an identifiable cadastral number in evidence,
        # expose a direct link to the public kadastrova-karta viewer.
        cadastral = f.evidence.get("cadastral")
        url = cadastral_url(cadastral)
        if url:
            ws.write_url(r, 9, url, url_fmt, str(cadastral))
        elif f.evidence.get("sample_cadastrals"):
            samples = f.evidence["sample_cadastrals"]
            first = next((c for c in samples if cadastral_url(c)), None)
            if first:
                ws.write_url(r, 9, cadastral_url(first), url_fmt,
                             f"{first}" + (f" (+{len(samples) - 1})" if len(samples) > 1 else ""))
            else:
                ws.write_blank(r, 9, None, wrap_fmt)
        else:
            ws.write_blank(r, 9, None, wrap_fmt)

    ws.freeze_panes(1, 0)
    if findings:
        ws.autofilter(0, 0, len(findings), len(headers) - 1)


def _short(v: Any) -> str:
    s = str(v)
    return s if len(s) <= 40 else s[:37] + "…"


def write_workbook(
    *,
    owners: list[list[Any]],
    owners_headers: list[str],
    owners_widths: list[int],
    owners_text_columns: list[int] | None = None,
    owners_date_columns: list[int] | None = None,
    owners_url_columns: dict[int, Callable[[Any], str | None]] | None = None,
    land: list[list[Any]],
    land_headers: list[str],
    land_widths: list[int],
    land_text_columns: list[int] | None = None,
    land_date_columns: list[int] | None = None,
    land_url_columns: dict[int, Callable[[Any], str | None]] | None = None,
    land_changed_masks: list[list[bool]] | None = None,
    realestate: list[list[Any]],
    re_headers: list[str],
    re_widths: list[int],
    re_text_columns: list[int] | None = None,
    re_date_columns: list[int] | None = None,
    re_url_columns: dict[int, Callable[[Any], str | None]] | None = None,
    re_changed_masks: list[list[bool]] | None = None,
    findings: list[Any] | None = None,
    metadata: dict[str, Any] | None = None,
) -> bytes:
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(
        buf,
        {
            "in_memory": True,           # don't touch disk
            "default_date_format": "yyyy-mm-dd",
            "strings_to_numbers": False,
            "strings_to_formulas": False,
            "strings_to_urls": False,
        },
    )

    formats = _build_formats(wb)

    _write_sheet(
        wb, "Реєстр власників", owners_headers, owners, formats,
        widths=owners_widths,
        text_columns=owners_text_columns,
        date_columns=owners_date_columns,
        url_columns=owners_url_columns,
    )
    _write_sheet(
        wb, "Земельні ділянки", land_headers, land, formats,
        widths=land_widths,
        text_columns=land_text_columns,
        date_columns=land_date_columns,
        url_columns=land_url_columns,
        changed_masks=land_changed_masks,
    )
    _write_sheet(
        wb, "Нерухомість", re_headers, realestate, formats,
        widths=re_widths,
        text_columns=re_text_columns,
        date_columns=re_date_columns,
        url_columns=re_url_columns,
        changed_masks=re_changed_masks,
    )

    if findings is not None:
        _write_findings_sheet(wb, formats, findings)

    if metadata:
        _write_metadata_sheet(wb, metadata, formats)

    wb.close()
    return buf.getvalue()
